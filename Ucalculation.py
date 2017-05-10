#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""


"""


import numpy as np
import openpyxl



def main():
    #first let's get all the necessary filtering done in isofilter
    
    spike_input = raw_input("What spike did you use? Options: DIII-B, DIII-A, 1I, 1H : ") #could use input files rather than raw_input
    printing =raw_input("Would you like to print as you go? [y/n] : ")
    correct_response  = 'yn'
    if printing.lower() not in correct_response:
        print ("yes or no please!")
    else:
        printing = printing.lower()
        
    AS_238_237_input = raw_input("What is your abundant sensitivity of 238U - 237U ? ")
    AS_229_230_input = raw_input("what is your abundant sensitivity of 229U - 230U ? ")
    #blank = raw_input("blank?(y/n): ")
    #if blank.lower()== y:
    #    blank_switch = 'yes'
    #    input_blank = raw_input("blank value?  ")
    #elif blank.lower() == n:
    #    blank_switch = 'no'
        
        
        
    spike = str(spike_input)
    spike_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
    if  spike_input in spike_dictionary:
        spike = spike_dictionary[spike]
    else:
        print ('no valid spike was entered')
    filename_U = raw_input("Enter the source file name to filter for U: ")
    filename_Th = raw_input("Enter the file name to filter for Th: ")
    '''
    u_wash_233 = input("What's your 233U wash mean value: ?(cps) ")
    u_wash_234 = input("What's your 234U wash mean value: ?(cps) ")
    u_wash_235 = input("What's your 235U wash mean value: ?(cps) ")
    th_wash_230 = input("What's your 230Th wash mean value: ?(cps)")
    darknoise = th_wash_230* 60
    u_blank_238 = input("What's your 238U blank value: ?(pmol)")
    th_blank_232 = input("What's your 232Th blank value: ?(pmol)")
    th_blank_230 = input("What's your 230Th blank value: ?(fmol)")
    
    
    '''    
    #filename_Uwash = raw_input("Enter the file name for U wash: ")
    #filename_Thwash = raw_input("Enter the file name for Th wash:")
    
    
    calc = Ucalculation(spike, printing,filename_U,filename_Th, AS_238_237_input, AS_229_230_input)
  
      
    dictTh = calc.U_normalization_forTh()
    dictU = calc.U_normalization_forAge()
    lisTh = calc.Th_calculation_for_age()
    

  
    
    if printing == 'y':
        
            print ("Results for Th filtering")
            print ("236/233 measured ratio: " + str(dictTh['236/233m']))
            print ("236/233 measured 2s error: " + str(dictTh['236/233error']))
            print ("235/233 corrected & normalized ratio: " + str(dictTh['235/233n']))
            print ("235/233 relative error: " + str(calc.five_three_err_rel))
            print ("235/233 corrected & normalized 2s error: " + str(calc.five_three_corrnorm_err))
            print ("236/233 corrected ratio: " + str(dictTh['236/233c']))
            print ("236/233 corrected 2s error: " + str(dictTh['236/233corr_err']))
            print ("Results for Age Calc")
            print ("235/233 corrected and normalized ratio : " + str(dictU['235/233n']))
            print ("235/233 corrected and normalized 2s error : " + str(dictU['235/233corrnorm_err']))
            print ("235/234 normalized and corrected ratio: " + str(dictU['235/234normcorr']))
            print ("235/234 normalized and corrected 2s error : " + str(dictU['235/234normcorr_err']))
            print ("Cycles of 233: " + str(dictU['233counts']))
            print ("Cycles of 234/235: " + str(dictU['234/235counts']))
            print ("Mean 233U cps : " + str(dictU['233mean']))
            print("Cell B18: "+str(lisTh[0]))
            print("Cell B17: "+ str(lisTh[1]))
            print("Cell C17: "+str(lisTh[2]))
            print("Cell C18: "+str(lisTh[3]))
            print("Cell C14: "+str(lisTh[4]))
            print("Cell C15: "+str(lisTh[5]))
            print("Cell B14: "+str(lisTh[6]))
            print("Cell B15: "+str(lisTh[7]))
            print("Cell B30: "+str(lisTh[8]))
            print("Cell B11: "+str(lisTh[9]))
            print("Cell O1002: "+str(lisTh[10]))
            print("Cell B12: "+str(lisTh[11]))
            print("Cell B10: "+str(lisTh[12]))
            print("Cell D11: " +str(lisTh[13]))
            print("Cell D12: " + str(lisTh[14]))
            print("Cell D10: " + str(lisTh[15]))
            print("Cell C10: " + str(lisTh[16]))
            print("Cell B6: " + str(lisTh[17]))
            print("cell B4: " +str(lisTh[18]))
            print("cell B20: " + str(lisTh[19]))
            print("Cell H21: " + str(lisTh[20]))
               
              
            
    
    
class UFilter():
    def __init__(self, filename,columnletter,filternumber): # input filename and columnletter as strings
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.filternumber = int(filternumber)
        self.workbook = openpyxl.load_workbook(self.filename)
        self.ws = self.workbook.active
        self.totalCounts = 0
        #self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
        self.filteredMean = 0
        self.filteredCounts = 0
        self.filteredStanddev = 0 
        self.filteredError = 0
        
    
        #All the get*** methods are unfiltered
    def getCounts(self):
        total_counts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row -8)):
            for cell in row:
                value = cell.value
                if value:
                    total_counts += 1

        self.totalCounts = total_counts
        return self.totalCounts
        
    def getMean(self):
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                if cell.value:
                    outlist.append(cell.value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        return self.mean
    
    def getStanddev(self):
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    outlist.append(value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        return self.standdev
        
    def filteredDict(self):
        outlist = []
        total_counts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    outlist.append(value)
                    total_counts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        self.totalCounts = total_counts
        self.criteria = self.filternumber * (self.standdev / self.totalCounts ** 0.5)
        outlist2 = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist2.append(np.nan)
                    else:
                        outlist2.append(value)
                        outcounts += 1
                else:
                    outlist2.append(np.nan)
        outarray2 = np.array(outlist2, dtype = np.float) 
        self.filteredMean = np.nanmean(a = outarray2 )
        self.filteredCounts= outcounts
        self.filteredStanddev = np.nanstd(a = outarray2, ddof = 1)
        self.filteredError = 2 * (self.filteredStanddev / (self.filteredCounts ** 0.5))
        
        
        dict1 = {'filteredMean': self.filteredMean,'filteredCounts':self.filteredCounts,\
        'filteredStanddev': self.filteredStanddev,'filteredError':self.filteredError}
        return dict1
        
        
        '''               
class ThFilter(self):
    def __init__(self, filename, columnletter, filternumber = 28):
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.filternumber = filternumber
        self.workbook = openpyxl.load_workbook(self.filename)
        self.ws = self.workbook.active
        self.totalCounts = 0
        #self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
        self.filteredMean = 0
        self.filteredCounts = 0
        self.filteredStanddev = 0 
        self.filteredError = 0
        
    def filteredDict():
        outlist = []
        total_counts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    outlist.append(value)
                    total_counts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        self.totalCounts = total_counts
        self.criteria = self.filternumber * (self.standdev / self.totalCounts ** 0.5)
        outlist2 = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist2.append(np.nan)
                    else:
                        outlist2.append(value)
                        outcounts += 1
                else:
                    outlist2.append(np.nan)
        outarray2 = np.array(outlist2, dtype = np.float) 
        self.filteredMean = np.nanmean(a = outarray2 )
        self.filteredCounts= outcounts
        self.filteredStanddev = np.nanstd(a = outarray2, ddof = 1)
        self.filteredError = 2 * (self.filteredStanddev / (self.filteredCounts ** 0.5))
        
        
               
    '''
class Ucalculation():
    """
    Class Ucalculation...
    """

    def __init__ (self,spike,printing,filename_U,filename_Th,abundent_sensitivity_238_237, abundant_sensitivity_229_230 ):
        
       
        self.spike = spike
        self.printing = printing
        self.filename_U = filename_U
        self.filename_Th = filename_Th
        self.wt_232 = 232.038051
        self.wt_235 = 235.043924
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_234 = 234.040947
        self.wt_230 = 230.033128
        self.wt_229 = 229.031756
        self.eight_five_raio = 137.83
        self.AS_229_230 = float(abundant_sensitivity_229_230)
        self.AS_238_237 = float(abundent_sensitivity_238_237)
        self.AS_236_238 = self.AS_238_237/5 #cell B28
        self.AS_234_238 = self.AS_238_237/20
        self.AS_232_230 = self.AS_229_230/5
        self.AS_232_229 = self.AS_232_230/3
        self.AS_238_236_b = self.AS_229_230/5
        
        
        #Filtering for 236/233
        filteredDict63 = UFilter(self.filename_U,"G", 44).filteredDict()
        self.six_three_mean_meas = filteredDict63['filteredMean']
        self.six_three_standdev = filteredDict63['filteredStanddev']
        self.six_three_err_meas = filteredDict63['filteredError']

        #Filtering for 235/233
        filteredDict53 =UFilter(self.filename_U, "H", 44).filteredDict()
        self.five_three_m = filteredDict53['filteredMean']
        self.five_three_standdev = filteredDict53['filteredStanddev']
        self.five_three_err_meas = filteredDict53['filteredError']
        
        #Filtering for 234/235
        filteredDict45 = UFilter(self.filename_U,"I", 44).filteredDict()
        self.four_five_mean_meas = filteredDict45['filteredMean']
        self.four_five_err_meas = filteredDict45['filteredError']
        self.four_five_counts = filteredDict45['filteredCounts']

        #Filtering for 229
        #Filtering for 230
        
        #Filtering for 230/229 column E

        #Not filtering 233
        self.three_mean_meas = UFilter(filename_U, "C", 44).getMean()
        self.three_counts = UFilter(filename_U, "C", 44).getCounts()
        
        #Filtering for 230/229 column E
        filteredDict09 =UFilter(self.filename_Th,"E", 28).filteredDict()
        self.zero_nine_m = filteredDict09['filteredMean']
        #self.zero_nine_m = UFilter(self.filename_Th,"E", 28).getMean()
        self.zero_nine_m_err= filteredDict09['filteredError']
        self.zero_nine_counts = filteredDict09['filteredCounts']

        #Filtering for 229/232 Column F
        
        filteredDict92 = UFilter(self.filename_Th,"F", 28).filteredDict()
        #self.nine_two_m = filteredDict92['filteredMean']
        self.nine_two_m = IsoFilter(self.filename_Th, "F", 28).getMean()
        self.nine_two_err_meas= filteredDict92['filteredError']
        self.nine_two_counts = filteredDict92['filteredCounts']
        self.nine_two_standdev = filteredDict92['filteredStanddev']

        #Filtering for 230/232 column G
        filteredDict02 = UFilter(self.filename_Th,"G", 28).filteredDict()
        #self.zero_two_m = filteredDict02['filteredMean']/1.02
        self.zero_two_m = UFilter(self.filename_Th,"G", 28).getMean()/1.02
        self.zero_two_err_meas= filteredDict02['filteredError']
        self.zero_two_counts = filteredDict02['filteredCounts']
        self.zero_two_standdev = filteredDict02['filteredStanddev']
        
        #Cell O1005 
       # self.std_229_232 = 2* (self.nine_two_standdev/(self.zero_two_counts**0.5))
        
        #not Filtering for 229Th column L
        self.nine_m = IsoFilter(self.filename_Th, "C", 28).getMean()
        self.nine_counts = IsoFilter(self.filename_Th,"C", 28).getCounts()

        self.U_wash_mean = 0
        self.Th_wash_mean = 0 
        
        self.six_three_c_err = 0
        self.six_three_corr_err_rel=0
        self.eight_five_rat_err_rel = 0.0003
        self.AS_six_eight_err_rel = 0.3
        self.AS_four_eight_err_rel = 0.3
        self.eight_five_rat_err_rel = 0.0003
        self.four_five_norm = 0
        self.six_three_c = 0
        #self.five_three_corrnorm_err =0
        self.five_three_err_rel = 0 
        #self.five_three_n = 100
        self.four_five_normcorr = 0
        self.four_five_normcorr_err = 0
        

    def U_normalization_forTh(self):
        #calculate (236/233)c
        self.six_three_c = self.six_three_mean_meas * \
        ( 1 - self.AS_236_238 * self.five_three_m * (self.eight_five_raio /self.spike) )
        
         #caulculate (235/233)n
        ratio = float((np.log(self.wt_235/self.wt_233))/np.log(self.wt_236/self.wt_233))
        self.five_three_n = self.five_three_m * (self.spike/self.six_three_c)**ratio

        #calculate the 235/233 relative error
        self.five_three_err_rel = self.five_three_err_meas/self.five_three_m
        
        #calculate the 236/233 relative error
        self.six_three_err_rel = self.six_three_err_meas/self.six_three_mean_meas
        
        #calculate the error of (236/233)c
        #self.six_three_corr_err = self.six_three_c * np.sqrt(self.six_three_err_rel**2 + \
        #s( ((0.3**2 * self.five_three_m * 137.83)/self.spike)* np.sqrt(self.AS_six_eight_err_rel**2 + self.five_three_err_rel ** 2 + self.eight_five_rat_err_rel**2 ))/ (1 - (self.AS_six_eight * self.five_three_m *137.83 / self.spike) ) ** 2 )
        
        #calculate the error of (236/233)c cell C6
        
        self.six_three_c_err = self.six_three_c * np.sqrt(self.six_three_err_rel**2 + \
        (((self.AS_236_238*self.five_three_m*137.83/self.spike)*np.sqrt(0.3**2+self.five_three_err_rel**2+(0.04/137.83)**2))/(1-self.AS_236_238*self.five_three_m*137.83/self.spike))**2)
                                                             
                                                             
        #calculate the relative error of (236/233)c  cellD6
        self.six_three_c_err_rel = self.six_three_c_err/ self.six_three_c
        
        #calculate error of (235/233)n cell c12
        self.five_three_corrnorm_err = self.five_three_n * \
        np.sqrt(self.five_three_err_rel**2 + (2 * self.six_three_corr_err_rel/3)**2)

        self.dictTh = {'236/233m': self.six_three_mean_meas,'236/233error':self.six_three_err_meas,'235/233n': self.five_three_n, \
                 '235/233corrnorm_err':self.five_three_corrnorm_err, '236/233c':self.six_three_c,'236/233corr_err':self.six_three_c_err }

        return self.dictTh
        
    def U_normalization_forAge(self):
        
        #calculate relative error of (234/235)m cell D10
        self.four_five_err_rel = self.four_five_err_meas / self.four_five_mean_meas
        ratio = float(np.log(self.wt_234/self.wt_235)/np.log(self.wt_236/self.wt_233))
        
        # calculate (234/235)n cell b13
        self.four_five_norm = self.four_five_mean_meas * (self.spike/self.six_three_c)**ratio

        #calculate error of (234/235)n cell c13
        self.four_five_norm_err = self.four_five_norm * np.sqrt(self.four_five_err_rel**2 + (ratio**2 * self.six_three_corr_err_rel**2) )

        #calculate the relative error of (234/235)n
        self.four_five_norm_err_rel = self.four_five_norm_err/self.four_five_norm
        
        #calculate the (234/235)nc cell B14
        self.four_five_normcorr = self.four_five_norm * (1 - (self.eight_five_raio * self.AS_234_238/ self.four_five_norm ))
        
        #calculate the error of (234/235)nc
        self.four_five_normcorr_err = self.four_five_normcorr * np.sqrt(self.four_five_norm_err_rel**2 +\
        ( (137.83 * self.AS_234_238 / self.four_five_norm) * np.sqrt(self.eight_five_rat_err_rel**2 + \
         self.AS_four_eight_err_rel**2 +self.four_five_norm_err_rel**2 )\
         / (1 - (137.83 * self.AS_234_238/ self.four_five_norm)) ) **2 )

       
        self.dictU = {'235/233n':self.five_three_n,'235/233corrnorm_err':self.five_three_corrnorm_err, '235/234normcorr':self.four_five_normcorr, \
                     '235/234normcorr_err':self.four_five_normcorr_err, '233counts':self.three_counts, '234/235counts':self.four_five_counts, '233mean':self.three_mean_meas}

        return self.dictU
        
    def Th_calculation_for_age(self):
        
        #B14 = self.zero_nine_c
        self.zero_nine_c = self.zero_nine_m *(1-self.AS_232_230/(self.zero_two_m)*(1-self.AS_229_230))
               
        #Calculating B6 
        self.six_three_c_Th = self.six_three_mean_meas *(1- self.AS_238_236_b* self.five_three_n*137.82/self.spike)
        
        #Calculating cell E20/B17  (230/229)cn
        ratio1 = float(np.log(self.wt_230/self.wt_229)/np.log(self.wt_236 / self.wt_233))
        self.zero_nine_cn = self.zero_nine_c *((self.spike)/(self.six_three_c_Th))**ratio1
        
        #Calculating 232/229_m cell B11
        self.two_nine_m = 1/(self.nine_two_m/1.02)
    
        #Calculating cell B15
        self.two_nine_c = self.two_nine_m * (1/(1-self.AS_232_229*self.two_nine_m))

        #calculating cell E21/B18 (232/229)cn
        ratio2 = float((np.log(self.wt_232/self.wt_229))/np.log(self.wt_236/self.wt_233))
        self.two_nine_cn = self.two_nine_c*((self.spike/self.six_three_c_Th)**ratio2)

        #calculating cell D10
        self.zero_nine_err_rel = self.zero_nine_m_err/self.zero_nine_m
        
        #Calculating cell D12 relative error or whatever it is 
        self.zero_two_m_err_rel = max(((2*(self.zero_two_standdev/(self.zero_two_counts**0.5)))/self.zero_two_m),0.02)

        #Calculating cell C14 (230/229)c error
        self.zero_nine_c_err = self.zero_nine_c \
        *np.sqrt(self.zero_nine_err_rel**2 + ((self.AS_232_230/self.zero_two_m)\
        *np.sqrt(0.3**2 +self.zero_two_m_err_rel **2)/(1-self.AS_232_230/self.zero_two_m))**2+\
        (self.AS_229_230*0.3/(1-self.AS_229_230))**2)
        
        #Calculating D14
        self.zero_nine_c_err_rel = self.zero_nine_c_err/self.zero_nine_c
        
        #Calculating D11 or whatever this is
        self.two_nine_err_rel = max(((2*(self.nine_two_standdev/(self.nine_two_counts**0.5)))/self.nine_two_m), 0.02)
        
        
        #Calculating cell C15
        self.two_nine_c_err = self.two_nine_c * np.sqrt(self.two_nine_err_rel**2 +(np.sqrt(0.3**2+ self.two_nine_err_rel**2)*(self.AS_232_229*self.two_nine_m)/(1-self.AS_232_229*self.two_nine_m)**2)**2)
        
        
        #Calculating D15 
        self.two_nine_c_err_rel = self.two_nine_c_err/self.two_nine_c
        
        #Calculating F20/C17 230/229cn error
        self.zero_nine_cn_err = self.zero_nine_cn* np.sqrt(self.zero_nine_c_err_rel**2+(self.six_three_c_err_rel/3)**2)
       
        #Calculating C11
        self.two_nine_m_err = self.two_nine_m * self.two_nine_err_rel
        
        
        #Calculating F21/C18 
        self.two_nine_cn_err = self.two_nine_cn* np.sqrt(self.two_nine_c_err_rel**2+self.six_three_c_err_rel**2)
        
       
        self.lis1 = [self.two_nine_cn, self.zero_nine_cn, self.zero_nine_cn_err,\
                     self.two_nine_cn_err,self.zero_nine_c_err, self.two_nine_c_err, self.zero_nine_c, \
                     self.two_nine_c, self.AS_232_229,self.two_nine_m,self.nine_two_m,self.zero_two_m, \
                     self.zero_nine_m, self.two_nine_err_rel, self.zero_two_m_err_rel, self.zero_nine_err_rel\
                     , self.zero_nine_m_err, self.six_three_c_Th, self.five_three_n,self.nine_m, self.nine_counts]
                     
        return self.lis1
        
        def Age_calculation(self):
            
            
            return

       
        
    
    
            
